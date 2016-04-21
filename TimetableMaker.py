from itertools import combinations

# Recebe dicionario de aulas de estrutura:
# { CLASS_ID:{
#       T:{
#           CLASS_T1,
#           CLASS_T...
#       },
#       TP:{
#           CLASS_TP...
#       },
#       ...
#   }, ...
# }

# Devolve array de todas as combinacoes de turmas possiveis
# Refere-se aos elementos no dicionario por tuples: (aula, tipo, turma)
# IGNORA SOBREPOSICOES
def possibleCombinations(dictionary):
    # Combinacoes de turmas validos (todos os tipos presentes)
    # Para cada aula, i.e., e necessario fazer combinacao de turmas
    combTurmasValidas = []

    aulas = []  # List de todas as aulas por numero

    #Fazer combinacoes dentro de cada aula
    for aula in dictionary:
        turmas = [] # List de de todas as turmas nesta aula (disciplina), como tuple
        tipos = []  # Tipos de aula (T/TP/PL)

        for tipo in dictionary[aula]:
            tipos.append(tipo)
            for turma in dictionary[aula][tipo]:
                turmas.append((aula, tipo, turma))

        combTurmas = combinations(turmas, len(tipos))   # Todas as combinacoes possiveis, incluindo (TP,TP,TP,TP)

        for comb in combTurmas:
            tiposNaComb = []   # Quais os tipos de aula nesta combinacao; deverao ser todos
            for turma in comb:
                tipo = turma[1] # Cada turma é representada por uma tuple (aula, tipo, turma); turma[1] devolve tipo

                if tipo not in tiposNaComb:
                    tiposNaComb.append(tipo)

            #Se a combinacao nao possuir todos os tipos de aula nao e valida
            if set(tiposNaComb) != set(tipos):
                continue

            combTurmasValidas.append(comb)

        aulas.append(aula)

    # Fazer combinacoes de aulas, tendo em conta combinacoes "legais" de turmas
    # Pelo mesmo processo que para as aulas:
    # Fazer todas as combinacoes possiveis e remover as que nao incluirem todas as aulas
    combAulas = combinations(combTurmasValidas, len(aulas))

    combAulasValidas = []   # Todas as combinacoes de turmas

    for comb in combAulas:
        aulasInComb = []    # List de aulas incluidas nesta combinacao; deverao ser todas
        for turmaComb in comb:  # Combinacao de turmas para uma aula; tira-se o id da aula pelo primeiro elemento
            if turmaComb[0][0] not in aulasInComb:
                aulasInComb.append(turmaComb[0][0])    # comb[0] == (aula, tipo, turma); tuple[0] == aula

        # Se esta combinacao de turmas possuir nao todas as aulas, nao e valida
        if set(aulasInComb) != set(aulas):
            continue

        # Verificar se a combinação nao existe ja sob outra ordem
        existe = False
        for combValida in combAulasValidas:
            if set(combValida) == set(comb):
                existe = True
                break
        if existe:
            continue

        combAulasValidas.append(comb)

    return combAulasValidas



# Recebe input:
    # Dicionario:
    # { CLASS_ID:{
    #       T:{
    #           [T1_obj, T1_obj, T1_obj,...],
    #           CLASS_T...
    #       },
    #       TP:{
    #           CLASS_TP...
    #       },
    #       ...
    #   }, ...

    # Combinacoes validas de aulas:
    # ( ( ((aula1, tipo1, turma1), (aula1, tipo2, turma1)), ((aula2, tipo1, turma1), (aula2, tipo2, turma1)) ), ... )

# Verifica se ha sobreposicao de aulas e se houver, remove-as
# Devolve lista de combinacoes sem sobreposicoes
def removeOverlaps(dictionary, validCombinations):
    noOverlaps = [] # Resultado a devolver

    for comb in validCombinations:
        turmas = [] # turmas com "coordenadas", sob a forma (horaInicio, horaFim, (aula, tipo, turma))

        # Criar tuples de horas e colocar na array
        for aulaComb in comb:
            for turma in aulaComb:
                aulas = dictionary[turma[0]][turma[1]][turma[2]] # Tirar objetos Aula do dicionario (multiplos!)

                for aula in aulas:
                    # Criar tuple com horas inicio/fim, dia e turma (disciplina/tipo/turma)
                    ref = (aula.horaInicio, aula.horaFim, aula.dia, (turma[0], turma[1], turma[2]))
                    turmas.append(ref)

        # Criar pares
        todosPares = combinations(turmas, 2)
        pares = []

        # Retirar pares de mesmas aulas
        for par in todosPares:
            # Verificar se turmas diferentes
            turmaA = par[0][3]
            turmaB = par[1][3]
            if turmaA[0] != turmaB[0] or turmaA[1] != turmaB[1] or turmaA[2] != turmaB[2]:
                pares.append(par)

        # Verificar sobreposicao em cada par
        combSemSobreposicoes = True

        for par in pares:
            a = par[0]
            b = par[1]

            # Dias diferentes?
            if a[2] != b[2]:
                continue

            cedo = min(a[0], b[0])
            tarde = max(a[1], b[1])
            delta = tarde - cedo

            # Aulas sobrepoem-se
            if a[1]-a[0]+b[1]-b[0] > delta:
                combSemSobreposicoes = False
                break

        if combSemSobreposicoes:
            noOverlaps.append(comb)


    return noOverlaps


from openpyxl import Workbook
from openpyxl.styles import Color, PatternFill, Style, Fill, Font
from random import randint

# Recebe input:
    # Dicionario:
    # { CLASS_ID:{
    #       T:{
    #           [T1_obj, T1_obj, T1_obj,...],
    #           CLASS_T...
    #       },
    #       TP:{
    #           CLASS_TP...
    #       },
    #       ...
    #   }, ...

    # Combinacoes de aulas:
    # ( ( ((aula1, tipo1, turma1), (aula1, tipo2, turma1)), ((aula2, tipo1, turma1), (aula2, tipo2, turma1)) ), ... )
# Grava um ficheiro xlsm (output.xlsm)
# Devolve workbook do openpyxl
def outputExcel(dictionary, combinations):
    if len(combinations) == 0:
        print("No combinations!")
        return

    wb = Workbook()
    wb.remove_sheet(wb.active) # Apagar folha default

    combinationNumber = 0
    for comb in combinations:
        ws = wb.create_sheet(str(combinationNumber))    # Criar uma nova folha com um id para referencia

        # Labels de dia
        ws['B1'] = "Segunda"
        ws['C1'] = "Terça"
        ws['D1'] = "Quarta"
        ws['E1'] = "Quinta"
        ws['F1'] = "Sexta"
        ws['G1'] = "Sabado"
        ws['H1'] = "Domingo"

        # Labels de hora (30/30 minutos, das 8 as 22)
        i = 2
        for n in range(80,220,5):
            ws['A'+str(i)] = str(int(n/10)) + "h" + str(int(((n/10)%1)*60)) + "m"
            i += 1

        # Desenhar aulas
        for disciplina in comb:
            for coord in disciplina:
                aulaObjList = dictionary[coord[0]][coord[1]][coord[2]]
                for aulaObj in aulaObjList:
                    # Tirar meia hora ao fim, para que nao haja merge sobreposto
                    cellRange = diaParaLetra(aulaObj.dia) + horaParaNumero(aulaObj.horaInicio) + ":"\
                           + diaParaLetra(aulaObj.dia) + horaParaNumero(aulaObj.horaFim - 0.5)

                    ws.merge_cells(cellRange)

                    # Add label
                    ws[diaParaLetra(aulaObj.dia) + horaParaNumero(aulaObj.horaInicio)] = aulaObj.aulaNome +\
                        "," + aulaObj.turma


        combinationNumber += 1  # Para referencia

    wb.save('output.xlsx')

    return wb


# ______ Helper functions para output: _________
def diaParaLetra(dia):
    if dia == 0:
        return "B"
    if dia == 1:
        return "C"
    if dia == 2:
        return "D"
    if dia == 3:
        return "E"
    if dia == 4:
        return "F"
    if dia == 5:
        return "G"
    if dia == 6:
        return "H"

def horaParaNumero(hora):
    delta = hora - 8
    return str(int(delta/0.5) + 2)
# _____________________________________________


# XLSXtoHTMLdemo

# Program to convert the data from an XLSX file to HTML.
# Uses the openpyxl library.

# Author: Vasudev Ram - http://www.dancingbison.com
# Altered by Miguel Murça for the purposes of this program

import openpyxl
from openpyxl import load_workbook

def convertExcelToWeb(workbook):
    worksheets = workbook._sheets

    for worksheet in worksheets:
        html_data = """
        <html>
            <head>
                <title>
                Horario
                </title>
            <head>
            <body>
            <table>
        """

        ws_range = worksheet.iter_rows('A1:I30')
        for row in ws_range:
            html_data += "<tr>"
            for cell in row:
                if cell.value is None:
                    html_data += "<td>" + ' ' + "<td>"
                else:
                    html_data += "<td>" + str(cell.value) + "<td>"
            html_data += "<tr>"
        html_data += "</table>\n</body>\n</html>"

        with open(worksheet.title + ".html", "w") as html_fil:
            html_fil.write(html_data)

    # EOF